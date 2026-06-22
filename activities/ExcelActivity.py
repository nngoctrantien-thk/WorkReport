from openpyxl import load_workbook
import pandas as pd

from activities.BaseActivity import BaseActivity


class ExcelActivity(BaseActivity):

    NAME = "ExcelActivity"

    DESCRIPTION = """
    Đọc dữ liệu từ file Excel (.xlsx).
    Có thể đọc một ô, một vùng dữ liệu hoặc toàn bộ sheet.
    """

    PARAMETERS = {
        "file_path": {
            "type": "string",
            "description": "Đường dẫn file Excel."
        },
        "action": {
            "type": "string",
            "description": "read_cell | read_range"
        },
        "sheet_name": {
            "type": "string",
            "description": "Tên sheet."
        },
        "cell": {
            "type": "string",
            "required": False
        },
        "cell_range": {
            "type": "string",
            "required": False
        },
        "add_header": {
            "type": "boolean",
            "required": False,
            "default": False
        }
    }

    EXAMPLES = [
        "Đọc ô A1 trong Sheet1",
        "Đọc vùng A1:D20 trong Sheet1",
        "Đọc toàn bộ Sheet1"
    ]

    @staticmethod
    def execute(
        file_path,
        action,
        sheet_name,
        cell=None,
        cell_range=None,
        add_header=False
    ):
        workbook = load_workbook(file_path)

        if sheet_name not in workbook.sheetnames:
            raise Exception(f"Sheet '{sheet_name}' không tồn tại.")

        sheet = workbook[sheet_name]

        if action == "read_cell":
            return sheet[cell].value

        if action == "read_range":
            data = []

            if cell_range:
                rows = sheet[cell_range]
            else:
                rows = sheet.iter_rows(values_only=True)

            for row in rows:
                if cell_range:
                    data.append([c.value for c in row])
                else:
                    data.append(list(row))

            if add_header:
                if not data:
                    return []

                return pd.DataFrame(
                    data[1:],
                    columns=data[0]
                ).to_dict("records")

            return data

        raise Exception(f"Unsupported action: {action}")