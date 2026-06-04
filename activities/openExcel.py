from openpyxl import load_workbook
import pandas as pd


class ExcelActivity:

    def __init__(self, file_path):
        self.workbook = load_workbook(file_path)

    def read_range(self, sheet_name, cell_range=None, add_header=False):
        sheet = self.workbook[sheet_name]

        data = []

        if not cell_range:
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
        else:
            for row in sheet[cell_range]:
                data.append([cell.value for cell in row])

        if add_header:
            if not data:
                return pd.DataFrame()

            return pd.DataFrame(
                data[1:],
                columns=data[0]
            ).to_dict("records")

        return data

    def read_cell(self, sheet_name, cell):
        sheet = self.workbook[sheet_name]
        return sheet[cell].value